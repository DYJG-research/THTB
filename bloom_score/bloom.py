# -*- coding: utf-8 -*-
import os
import json
import traceback
import logging
import time
from openai import OpenAI
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from tqdm import tqdm
from typing import Any, Optional
from pathlib import Path
import ast
import operator
import math
import argparse


ALLOWED_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}

ALLOWED_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def _safe_eval_arithmetic(expression: str) -> float:
    """
    Safely evaluate arithmetic expressions containing only + - * / and parentheses.

    Only allows numbers, parentheses, and unary/binary addition, subtraction, multiplication, division.
    """
    try:
        parsed = ast.parse(expression, mode="eval")
    except SyntaxError as exc:  # noqa: BLE001
        raise ValueError(f"Invalid expression: {expression}") from exc

    def _eval(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Num):  # Compatibility with older Python versions
            return float(node.n)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return float(node.value)
            raise ValueError("Unsupported constant type")
        if isinstance(node, ast.BinOp):
            if type(node.op) not in ALLOWED_BINARY_OPERATORS:  
                raise ValueError("Unsupported binary operator")
            left = _eval(node.left)
            right = _eval(node.right)
            op_func = ALLOWED_BINARY_OPERATORS[type(node.op)]
            return op_func(left, right)
        if isinstance(node, ast.UnaryOp):
            if type(node.op) not in ALLOWED_UNARY_OPERATORS:  
                raise ValueError("Unsupported unary operator")
            operand = _eval(node.operand)
            op_func = ALLOWED_UNARY_OPERATORS[type(node.op)]
            return op_func(operand)
        raise ValueError("Contains disallowed syntax nodes")

    value = _eval(parsed)
    if not math.isfinite(value):
        raise ValueError("Expression evaluation result is invalid")
    return value


_NUMBER_REGEX = re.compile(r"[-+]?\d+(?:\.\d+)?")
_ALLOWED_CHARS_REGEX = re.compile(r"[^0-9+\-*/().]")


def _normalize_text(text: str) -> str:
    text = text.strip()
    text = (
        text.replace("，", ",")
        .replace("＋", "+")
        .replace("－", "-")
        .replace("×", "*")
        .replace("÷", "/")
        .replace("（", "(")
        .replace("）", ")")
        .replace("＝", "=")
    )
    return text


def extract_final_score(value: Any) -> Optional[float]:
    """
    Convert any form of score value to final numerical value:
    - If contains equals sign, prioritize taking the first number after the equals sign
    - Otherwise try to treat content as expression containing only +,-,*,/ and parentheses for safe evaluation
    - If both fail, fall back to extracting the first number from the full text
    Returns float or None (if cannot parse)
    """
    if pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, bool):
            return float(int(value))
        return float(value)

    s = _normalize_text(str(value))
    if s == "":
        return None

    if "=" in s:
        tail = s.rsplit("=", 1)[-1]
        m = _NUMBER_REGEX.search(tail)
        if m:
            try:
                return float(m.group(0))
            except Exception:  
                pass

    expr = _ALLOWED_CHARS_REGEX.sub("", s)
    if expr:
        try:
            result = _safe_eval_arithmetic(expr)
            return float(result)
        except Exception:  
            pass

    m2 = _NUMBER_REGEX.search(s)
    if m2:
        try:
            return float(m2.group(0))
        except Exception:  
            return None

    return None


def _find_score_column_name(columns: pd.Index) -> Optional[str]:
    for col in columns:
        if "score" in str(col).lower():
            return col
    return None


def clean_excel_score_column(file_path: Path, inplace: bool = False) -> None:
    """Clean score columns in Excel file, converting string expressions to numerical values."""
    
    def _to_int_if_close(x: Any) -> Any:
        if isinstance(x, float) and x.is_integer():
            return int(x)
        return x

    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        score_col = _find_score_column_name(df.columns)
        if score_col is None:
            continue
            
        print(f"Processing sheet '{sheet_name}', score column: '{score_col}'")
        
        cleaned_scores = df[score_col].apply(extract_final_score).apply(_to_int_if_close)
        
        score_col_idx = df.columns.get_loc(score_col) + 1  # Excel columns are 1-indexed
        
        for row_idx, cleaned_score in enumerate(cleaned_scores, start=2):  # Data starts from row 2
            ws.cell(row=row_idx, column=score_col_idx, value=cleaned_score)
    
    output_path = file_path if inplace else file_path.with_stem(f"{file_path.stem}_cleaned")
    wb.save(output_path)
    print(f"✅ Cleaned Excel saved to: {output_path}")


def add_normalized_bloom_score_column(file_path: Path) -> None:
    """Add a new column to each worksheet in Excel containing score column: normalized_bloom_score = (score - 1) / 20."""
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        score_col = _find_score_column_name(df.columns)
        if score_col is None:
            continue
            
        print(f"Adding normalized score column to sheet '{sheet_name}'")
        
        # Add header
        new_col_idx = len(df.columns) + 1
        ws.cell(row=1, column=new_col_idx, value="normalized_bloom_score")
        
        # Add normalized scores
        score_col_idx = df.columns.get_loc(score_col) + 1
        for row_idx in range(2, len(df) + 2):  # Data starts from row 2
            score_cell = ws.cell(row=row_idx, column=score_col_idx)
            if score_cell.value is not None:
                try:
                    normalized_score = (float(score_cell.value) - 1) / 20
                    ws.cell(row=row_idx, column=new_col_idx, value=normalized_score)
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=new_col_idx, value=None)
    
    wb.save(file_path)
    print(f"✅ Normalized score column added to: {file_path}")


def optimize_answer(input_path, output_path):
    """Main function to process Bloom taxonomy classification and scoring."""
    
    # Try to get API configuration from environment variables first
    api_key = os.getenv('OPENAI_API_KEY')
    api_base = os.getenv('OPENAI_API_BASE')
    model_name = os.getenv('OPENAI_MODEL')
    
    # If environment variables not found, try to load from config file
    if not api_key or not api_base or not model_name:
        try:
            # Look for config file in parent directory
            script_dir = Path(__file__).parent
            config_path = script_dir.parent / "config_template.json"
            
            if config_path.exists():
                print(f"Loading API configuration from: {config_path}")
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                llm_config = config.get('llm_api_config', {})
                api_key = llm_config.get('api_key', 'EMPTY')
                api_base = llm_config.get('api_base')
                model_name = llm_config.get('model')
                
                print(f"Loaded API config - Base: {api_base}, Model: {model_name}")
            else:
                print(f"Config file not found at: {config_path}")
        except Exception as e:
            print(f"Failed to load config file: {e}")
    
    if not api_key or not api_base or not model_name:
        raise ValueError("API configuration not found. Please set OPENAI_API_KEY, OPENAI_API_BASE, and OPENAI_MODEL environment variables, or ensure config_template.json exists with llm_api_config section.")
    
    client = OpenAI(
        api_key=api_key,
        base_url=api_base
    )

    # Configure logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    # Load input Excel file
    input_wb = load_workbook(input_path)
    input_sheet = input_wb.active

    # Create output Excel file
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Bloom Classification Results"

    # Set headers (use lowercase to match pipeline expectations)
    headers = ["query", "response", "Bloom Level", "Score"]
    output_sheet.append(headers)

    # Get data from input sheet
    data_rows = []
    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Ensure both query and response exist
            data_rows.append((row[0], row[1]))

    print(f"Found {len(data_rows)} data rows to process")

    # Process each row
    max_retries = 3
    retry_delay = 2

    with tqdm(total=len(data_rows), desc="Processing Bloom Classification") as pbar:
        for query_context, response_context in data_rows:
            
            # Create cells for level and score
            level_cell = openpyxl.cell.Cell(output_sheet)
            score_cell = openpyxl.cell.Cell(output_sheet)

            # Construct prompt
            prompt = f"""\\no_think Please analyze the following educational question-answer pair and classify it according to Bloom's Taxonomy. Provide your analysis in JSON format.

Question: {query_context}
Answer: {response_context}

Please classify this Q&A pair according to Bloom's Taxonomy and provide a difficulty score from 1-21:

**Bloom's Taxonomy Levels:**
1. **Remember (1-3 points)**: Recall facts, basic concepts, answers
   - Examples: Define, list, identify, name, state
   - Characteristics: Direct recall, memorization, basic factual knowledge

2. **Understand (4-6 points)**: Explain ideas or concepts, interpret meaning
   - Examples: Explain, describe, summarize, interpret, classify
   - Characteristics: Comprehension, explanation in own words, basic interpretation

3. **Apply (7-9 points)**: Use information in new situations, solve problems using acquired knowledge
   - Examples: Calculate, solve, demonstrate, apply, use
   - Characteristics: Practical application, using knowledge in new contexts

4. **Analyze (10-12 points)**: Draw connections among ideas, examine and break down information
   - Examples: Compare, contrast, examine, analyze, categorize
   - Characteristics: Breaking down complex information, identifying relationships

5. **Evaluate (13-15 points)**: Justify a stand or decision, critique based on standards
   - Examples: Judge, critique, evaluate, assess, argue
   - Characteristics: Making judgments, critical assessment, defending positions

6. **Create (16-21 points)**: Produce new or original work, combine elements in new ways
   - Examples: Design, construct, create, develop, formulate
   - Characteristics: Original thinking, synthesis, innovation, generating new ideas

**Scoring Guidelines:**
- Within each level, assign scores based on complexity and depth
- Consider the cognitive demand required to answer the question
- Higher scores within a level indicate greater complexity or multiple sub-skills

Please respond in the following JSON format:
{{
    "level": "Level Name (e.g., 'Apply')",
    "score": numerical_score_1_to_21,
    "reasoning": "Brief explanation of classification and scoring rationale"
}}"""

            # Retry mechanism for API calls
            for retry in range(max_retries):
                try:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "user", "content": prompt}
                        ],
                        stream=True,  # Enable streaming output
                    )

                    full_response = ""
                    for chunk in response:
                        if chunk.choices[0].delta.content:  # Check if there's new content
                            chunk_content = chunk.choices[0].delta.content
                            print(chunk_content, end="", flush=True)  # Print in real-time
                            full_response += chunk_content  # Accumulate complete response

                    cleaned_response = re.sub(r'<think>.*?</think>', '', full_response, flags=re.DOTALL)
                    cleaned_response = cleaned_response.replace("```json", '')
                    cleaned_response = cleaned_response.replace("```", '')
                    cleaned_response = cleaned_response.strip()

                    logging.info("Model response:")
                    logging.info(cleaned_response)
                    print("Model response:")
                    print(cleaned_response)
                    resp_json = json.loads(cleaned_response)

                    level_cell.value = resp_json["level"]
                    level_part = level_cell.value

                    raw_score = resp_json["score"]
                    final_score = extract_final_score(raw_score)
                    if isinstance(final_score, float) and final_score.is_integer():
                        final_score = int(final_score)
                    score_cell.value = final_score if final_score is not None else raw_score
                    score_part = score_cell.value

                    new_row = [
                        query_context,  # Original question
                        response_context, # Original answer
                        level_part,
                        score_part,  # Cleaned score
                    ]
                    output_sheet.append(new_row)
                    output_wb.save(output_path)  # Save in real-time
                    print(f"✅ Appended and saved to {output_path}")

                    break  # Success, break out of retry loop

                except Exception as e:
                    if retry < max_retries - 1:
                        logging.warning(f"Request failed (retry {retry + 1}): {e}, waiting {retry_delay} seconds before retry...")
                        print(f"⚠️ Request failed (retry {retry + 1}): {e}, waiting {retry_delay} seconds before retry...")
                        time.sleep(retry_delay)
                    else:
                        logging.error(f"Request failed (reached maximum retries): {e}")
                        traceback.print_exc()
                        print(f"❌ Request failed (reached maximum retries): {e}")
            pbar.update(1)

    # After all data is processed, add normalized score column
    try:
        add_normalized_bloom_score_column(Path(output_path))
        print("✅ Added normalized score column 'normalized_bloom_score'")
    except Exception as e:
        logging.error(f"Failed to add normalized score column: {e}")
        print(f"⚠️ Failed to add normalized score column: {e}")


def cli_main() -> None:
    parser = argparse.ArgumentParser(description="Generate and clean Bloom scores, and add normalized score column")
    parser.add_argument("-i", "--input", required=True, help="Input Excel path, e.g.: train-alpaca_No_3.xlsx")
    parser.add_argument("-o", "--output", required=True, help="Output Excel path, e.g.: bloom_alpaca_classify_No_3.xlsx")
    args = parser.parse_args()
    optimize_answer(args.input, args.output)


if __name__ == "__main__":
    cli_main()
